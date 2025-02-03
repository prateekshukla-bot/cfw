from flask import Flask
from flask_mysqldb import MySQL
from openpyxl import load_workbook

app = Flask(__name__)

# MySQL configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''  # Default password for root user is empty
app.config['MYSQL_DB'] = 'jnayil_db'

# Initialize MySQL
mysql = MySQL(app)

@app.route('/')                                                                               
def import_data():
    # Path to the XLSX file you want to import
    file_path = "D:/USER PROFILE DATA/Downloads/invoice list 2 (3).XLSX"


    # Load the workbook and sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active  # Assuming the first sheet
    
    # Open MySQL connection
    cur = mysql.connection.cursor()

    # Create the table if it doesn't exist
    create_table_query = """
    CREATE TABLE IF NOT EXISTS invoices (
        id INT AUTO_INCREMENT PRIMARY KEY,
        invoice_number VARCHAR(255),
        invoice_item VARCHAR(255),
        invoice_date DATE,
        customer VARCHAR(255),
        customer_name VARCHAR(255),
        invoice_quantity INT,
        material VARCHAR(255),
        material_description TEXT,
        sales_unit VARCHAR(50)
    );
    """
    cur.execute(create_table_query)
    
    # Iterate over each row in the sheet and insert data into the database
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header
        # Ensure only 9 columns are extracted
        if len(row) == 9:
            invoice_number, invoice_item, invoice_date, customer, customer_name, invoice_quantity, material, material_description, sales_unit = row
            
            # Insert data into the invoices table
            insert_query = """
            INSERT INTO invoices (invoice_number, invoice_item, invoice_date, customer, customer_name, invoice_quantity, material, material_description, sales_unit)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s);
            """
            cur.execute(insert_query, (invoice_number, invoice_item, invoice_date, customer, customer_name, invoice_quantity, material, material_description, sales_unit))
        else:
            # Handle rows with unexpected column count (e.g., print them or skip)
            print(f"Skipping invalid row: {row}")
    
    # Commit the changes and close the cursor
    mysql.connection.commit()
    cur.close()

    return "Data imported successfully!"

if __name__ == '__main__':
    app.run(debug=True)